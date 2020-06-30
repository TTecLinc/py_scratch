# -*- coding: utf-8 -*-
"""
Created on Mon Jun 29 21:27:57 2020

@author: Peilin Yang
"""
 

import openpyxl

import os

 

def getfile():

    filenames = os.listdir('.')

    mf = []

    for filename in filenames:

        if 'xlsx' in filename:

            mf.append(filename)

    return mf

 

def copyxl(wb,f):

    print(f)

    wb1 = openpyxl.load_workbook(f)

    wb1names = wb1.sheetnames

    print(wb1names)

    wbnames = wb.sheetnames

    print(wbnames)

    for name in wb1names:

        if name not in wbnames:

            wb.create_sheet(name)

            ws = wb[name]

            maxr = 0

        else:

            ws = wb[name]

            maxr = ws.max_row

        

        ws1 = wb1[name]

        for m in range(1,ws1.max_row+1):

            for n in range(1,1+ws1.max_column):

                cell = ws1.cell(m,n).value

                ws.cell(maxr+m,n).value = cell

 

    wb.save('merge.xlsx')

    wb1.close()

                

 

def mergexl():

    wb =openpyxl.Workbook()

    filenames = getfile()

    for f in filenames:

        copyxl(wb,f)

    wb.close()

 

if __name__ == '__main__':

    mergexl()
